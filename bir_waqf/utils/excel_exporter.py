import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io, json
import frappe
from bir_waqf.utils.project_utils import get_project_title, resolve_project_tokens

def build_transactions_excel(transactions):
	"""
	Generates an openpyxl Workbook in-memory for selected Bir Transactions.
	Applies RTL direction, custom Arabic styling, and auto column width calculations.
	"""
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "المعاملات المحددة"

	ws.views.sheetView[0].rightToLeft = True

	title_fill = PatternFill(start_color="0A4D2E", end_color="0A4D2E", fill_type="solid")
	title_font = Font(name="Tajawal", size=14, bold=True, color="FFFFFF")

	header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
	header_font = Font(name="Tajawal", size=10, bold=True, color="FFFFFF")

	thin_border = Border(
		left=Side(style='thin', color='E2E8F0'),
		right=Side(style='thin', color='E2E8F0'),
		top=Side(style='thin', color='E2E8F0'),
		bottom=Side(style='thin', color='E2E8F0')
	)

	ws.merge_cells("A1:I1")
	ws["A1"] = "تقرير قائمة معاملات منصة البر الوقفية المحددة للمطابقة"
	ws["A1"].fill = title_fill
	ws["A1"].font = title_font
	ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
	ws.row_dimensions[1].height = 35

	ws.append([]) # empty row 2

	headers = [
		"#", "رقم المعاملة", "رقم الحوالة / الصك", "اسم المتبرع",
		"المصرف", "المشروع", "القيمة الإجمالية (د.ل)", "تاريخ المعاملة", "حالة المطابقة"
	]
	ws.append(headers)
	ws.row_dimensions[3].height = 25

	for col_num in range(1, len(headers) + 1):
		cell = ws.cell(row=3, column=col_num)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = thin_border

	for idx, tx_info in enumerate(transactions, 1):
		tx = frappe.get_doc("Bir Transaction", tx_info.get("name") or tx_info.get("transaction_id"))
		
		proj_display = "-"
		if tx.is_basket:
			sub_projs = tx.basket_projects or []
			if sub_projs:
				titles = [get_project_title(p.project_name) for p in sub_projs if p.project_name]
				proj_display = ", ".join(titles)
		else:
			if tx.project:
				proj_display = get_project_title(tx.project)

		dt_str = str(tx.transaction_date)[:16] if tx.transaction_date else "-"
		rec_status = tx.reconciliation_status or "غير مطابق"

		row_data = [
			idx,
			tx.transaction_id or "-",
			tx.transfer_number or "-",
			tx.donor_name or "فاعل خير",
			tx.bank_name or "-",
			proj_display,
			tx.total_amount or 0.0,
			dt_str,
			rec_status
		]
		ws.append(row_data)

		row_idx = ws.max_row
		ws.row_dimensions[row_idx].height = 20
		for col_num in range(1, len(headers) + 1):
			c = ws.cell(row=row_idx, column=col_num)
			c.border = thin_border
			c.alignment = Alignment(horizontal="right", vertical="center")
			if col_num in [1, 8, 9]:
				c.alignment = Alignment(horizontal="center", vertical="center")
			if col_num == 7:
				c.number_format = '#,##0.00 "د.ل"'
				c.font = Font(name="Tajawal", bold=True, color="0A4D2E")

	for col in ws.columns:
		max_len = 0
		col_letter = get_column_letter(col[0].column)
		for cell in col:
			val_str = str(cell.value or "")
			if cell.row > 1 and len(val_str) > max_len:
				max_len = len(val_str)
		ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

	output = io.BytesIO()
	wb.save(output)
	return output.getvalue()


def build_grouped_bank_statement_excel(import_batch, bank, selected_projects):
	"""
	Generates an openpyxl Workbook grouped by Project for Quick Entry / Bank Reconciliation.
	Displays Arabic Project Titles and resolves project tokens.
	"""
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "كشف حساب المصرف - المشاريع"

	ws.views.sheetView[0].rightToLeft = True

	title_fill = PatternFill(start_color="0A4D2E", end_color="0A4D2E", fill_type="solid")
	title_font = Font(name="Tajawal", size=13, bold=True, color="FFFFFF")
	
	proj_header_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
	proj_header_font = Font(name="Tajawal", size=11, bold=True, color="0A4D2E")
	
	subtotal_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
	subtotal_font = Font(name="Tajawal", size=11, bold=True, color="92400E")

	header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
	header_font = Font(name="Tajawal", size=10, bold=True, color="FFFFFF")

	thin_border = Border(
		left=Side(style='thin', color='CBD5E1'),
		right=Side(style='thin', color='CBD5E1'),
		top=Side(style='thin', color='CBD5E1'),
		bottom=Side(style='thin', color='CBD5E1')
	)

	# Title Banner
	ws.merge_cells("A1:G1")
	ws["A1"] = f"كشف الحساب وتوزيع التبرعات — المصرف: {bank or 'الكل'} (الدفعة: {import_batch or 'الكل'})"
	ws["A1"].fill = title_fill
	ws["A1"].font = title_font
	ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
	ws.row_dimensions[1].height = 32

	ws.append([]) # empty row 2

	headers = ["#", "رقم المعاملة", "رقم الحوالة / الصك", "المستخدم / المتبرع", "مبلغ التبرع (د.ل)", "تاريخ المعاملة", "تمت المطابقة"]
	ws.append(headers)
	ws.row_dimensions[3].height = 24
	for c_idx in range(1, len(headers) + 1):
		cell = ws.cell(row=3, column=c_idx)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = thin_border

	filters = {}
	if import_batch and str(import_batch).strip():
		filters["import_batch"] = str(import_batch).strip()
	if bank and str(bank).strip():
		filters["bank_name"] = str(bank).strip()

	if isinstance(selected_projects, str):
		if selected_projects.startswith("["):
			try:
				selected_projects = json.loads(selected_projects)
			except Exception:
				selected_projects = [p.strip() for p in selected_projects.split(",") if p.strip()]
		else:
			selected_projects = [p.strip() for p in selected_projects.split(",") if p.strip()]

	if not selected_projects:
		selected_projects = []

	cur_row = 4

	for p_input in selected_projects:
		if not p_input or not str(p_input).strip():
			continue

		clean_p = str(p_input).strip()
		p_title = get_project_title(clean_p)
		tokens = list(resolve_project_tokens(clean_p))

		if not tokens:
			tokens = [clean_p.lower()]

		# Query matching single and basket transactions using tokens
		txs_single = frappe.get_all(
			"Bir Transaction",
			filters={**filters, "is_basket": 0, "project": ["in", tokens]},
			fields=["name", "transaction_id", "transfer_number", "donor_name", "total_amount", "transaction_date", "reconciliation_status"]
		)

		placeholders = ", ".join(["%s"] * len(tokens))
		sql_query = f"""
			SELECT t.name, t.transaction_id, t.transfer_number, t.donor_name, b.sub_amount as total_amount, t.transaction_date, t.reconciliation_status
			FROM `tabBir Transaction` t
			INNER JOIN `tabBir Basket Project` b ON b.parent = t.name
			WHERE t.is_basket = 1
			{" AND t.import_batch = %s" if import_batch else ""}
			{" AND t.bank_name = %s" if bank else ""}
			AND LOWER(TRIM(b.project_name)) IN ({placeholders})
		"""
		params = [v for v in [import_batch, bank] if v] + tokens
		txs_basket_rows = frappe.db.sql(sql_query, tuple(params), as_dict=True) or []

		all_project_txs = txs_single + txs_basket_rows

		# Group Section Header with Arabic Title
		ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
		header_cell = ws.cell(row=cur_row, column=1)
		header_cell.value = f"📌 مشروع: {p_title} (عدد التبرعات: {len(all_project_txs)})"
		header_cell.fill = proj_header_fill
		header_cell.font = proj_header_font
		header_cell.alignment = Alignment(horizontal="right", vertical="center")
		ws.row_dimensions[cur_row].height = 24
		cur_row += 1

		proj_sum = 0.0
		for t_idx, tx in enumerate(all_project_txs, 1):
			amt = float(tx.total_amount or 0.0)
			proj_sum += amt
			is_reconciled = "نعم" if tx.reconciliation_status in ["مطابق آليًا", "مطابق يدويًا"] else "لا"
			dt_str = str(tx.transaction_date)[:16] if tx.transaction_date else "-"

			row_vals = [
				t_idx,
				tx.transaction_id or "-",
				tx.transfer_number or "-",
				tx.donor_name or "فاعل خير",
				amt,
				dt_str,
				is_reconciled
			]
			ws.append(row_vals)
			ws.cell(row=cur_row, column=5).number_format = '#,##0.00 "د.ل"'
			ws.cell(row=cur_row, column=7).alignment = Alignment(horizontal="center")
			
			for c_col in range(1, 8):
				ws.cell(row=cur_row, column=c_col).border = thin_border

			ws.row_dimensions[cur_row].height = 20
			cur_row += 1

		# Subtotal Row
		ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=4)
		ws.cell(row=cur_row, column=1).value = f"إجمالي تبرعات مشروع ({p_title}):"
		ws.cell(row=cur_row, column=1).fill = subtotal_fill
		ws.cell(row=cur_row, column=1).font = subtotal_font
		ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

		ws.cell(row=cur_row, column=5).value = proj_sum
		ws.cell(row=cur_row, column=5).fill = subtotal_fill
		ws.cell(row=cur_row, column=5).font = subtotal_font
		ws.cell(row=cur_row, column=5).number_format = '#,##0.00 "د.ل"'

		ws.merge_cells(start_row=cur_row, start_column=6, end_row=cur_row, end_column=7)
		ws.cell(row=cur_row, column=6).fill = subtotal_fill

		for c_col in range(1, 8):
			ws.cell(row=cur_row, column=c_col).border = thin_border

		ws.row_dimensions[cur_row].height = 22
		cur_row += 2

	for col in ws.columns:
		max_len = 0
		col_letter = get_column_letter(col[0].column)
		for cell in col:
			val_str = str(cell.value or "")
			if cell.row > 1 and len(val_str) > max_len:
				max_len = len(val_str)
		ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

	output = io.BytesIO()
	wb.save(output)
	return output.getvalue()
